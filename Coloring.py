from openpyxl.styles import PatternFill, Border, Side

# !!!Found a script for colors methods, allowing for easier identification!!!
def highlight_excel_sections(wb):
    # Define fill colors for each section
    section_fills = {
        "economy": PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),  # Light Yellow
        "connection": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Light Green
        "inventory": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),  # Light Blue
        "k/d": PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid"),  # Light Orange
        "player": PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid"),  # Light Purple
        "health": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # Light Red
    }

    # Define what keywords belong to each section
    column_section_map = {
        "connected": "connection",
        "ping": "connection",
        "active_weapon_original_owner": "player",
        "steamid": "player",
        "name": "player",
        "prev_owner": "player",
        "inventory_as_ids": "inventory",
        "inventory": "inventory",
        "item_def_idx": "inventory",
        "item_id_high": "inventory",
        "item_id_low": "inventory",
        "inventory_position": "inventory",
        "is_silencer_on": "inventory",
        "has_defuser": "inventory",
        "has_helmet": "inventory",
        "utility_damage_total": "k/d",
        "enemies_flashed_total": "k/d",
        "ace_rounds_total": "k/d",
        "4k_rounds_total": "k/d",
        "3k_rounds_total": "k/d",
        "score": "k/d",
        "mvps": "k/d",
        "kills_total": "k/d",
        "headshot_kills_total": "k/d",
        "balance": "economy",
        "start_balance": "economy",
        "total_cash_spent": "economy",
        "kill_reward_total": "economy",
        "cash_earned_total": "economy",
        "money_saved_total": "economy",
        "current_equip_value": "economy",
        "round_start_equip_value": "economy",
        "cash_spent_this_round": "economy",
        "alive_time_total": "health",
        "health": "health",
        "life_state": "health",
        "is_alive": "health",
        "armor_value": "health"
    }

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:  # Row 1 = headers
            header = str(cell.value).lower()
            matched_section = None
            for keyword, section in column_section_map.items():
                if keyword in header:
                    matched_section = section
                    break
            if matched_section:
                fill = section_fills[matched_section]
                cell.fill = fill