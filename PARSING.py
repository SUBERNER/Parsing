from demoparser2 import DemoParser  # https://github.com/LaihoE/demoparser
import pandas as pd
from openpyxl import load_workbook
import os
from Coloring import highlight_excel_sections # Found online for coloring excel files

def get_df(title: str, ticks):
    # initialize a list to store all round data
    data = []
    # loop through each tick and collect data
    print(title)
    for tick in ticks:
        print(f"tick: {tick}")
        df = parser.parse_ticks(fields, ticks=[tick])
        data.append(df)  # store the dataframe
    # combine all dataframes into a single one
    return pd.concat(data, ignore_index=True)


# PUT DEMO FILE HERE
# PUT DEMO FILE HERE
# PUT DEMO FILE HERE
files = [os.getcwd() + "\\DEMO\\auto-20250324-2324-de_dust2-Counter-Strike_2.dem"]

for file in files:
    parser = DemoParser(file)

    """
    # If you just want the names of all events then you can use this:
    event_names = parser.list_game_events()
    df = parser.parse_events(["all"])
    print(df)
    print(type(df))
    """

    header = parser.parse_header()
    print(f"\naddons: {header["addons"]}")
    print(f"version: {header["demo_version_name"]}")
    print(f"guid: {header["demo_version_guid"]}")
    print(f"server: {header["server_name"]}")
    print(f"client: {header["client_name"]}")
    print(f"map: {header["map_name"]}\n")

    # selecting data to collect
    #round_mvp_ticks = parser.parse_event("round_mvp")["tick"].tolist()
    player_death_ticks = parser.parse_event("player_death")["tick"].tolist()
    ##bomb_planted_ticks = parser.parse_event("bomb_planted")["tick"].tolist()
    ##bomb_defused_ticks = parser.parse_event("bomb_defused")["tick"].tolist()
    #hostage_rescued_ticks = parser.parse_event("hostage_rescued")["tick"].tolist()
    weapon_fire_ticks = parser.parse_event("weapon_fire")["tick"].tolist()
    #flashbang_detonate_ticks = parser.parse_event("flashbang_detonate")["tick"].tolist()
    #hegrenade_detonate_ticks = parser.parse_event("hegrenade_detonate")["tick"].tolist()
    #molotov_detonate_tick = parser.parse_event("molotov_detonate")["tick"].tolist()
    #smokegrenade_detonate_ticks = parser.parse_event("smokegrenade_detonate")["tick"].tolist()
    player_hurt_ticks = parser.parse_event("player_hurt")["tick"].tolist()
    ##player_blind_ticks = parser.parse_event("player_blind")["tick"].tolist()
    player_connect_ticks = parser.parse_event("player_connect")["tick"].tolist()
    #player_activate_ticks = parser.parse_event("player_activate")["tick"].tolist()
    begin_new_match_ticks = parser.parse_event("begin_new_match")["tick"].tolist()

    round_start_ticks = parser.parse_event("round_start")["tick"].tolist()
    round_freeze_end_ticks = parser.parse_event("round_freeze_end")["tick"].tolist()
    round_end_ticks = parser.parse_event("round_end")["tick"].tolist()

    fields = ["start_balance",
              "balance",
              "cash_spent_this_round",
              "total_cash_spent",
              "is_alive",
              "health",
              "armor_value",
              "life_state",
              "round_start_equip_value",
              "current_equip_value",
              "kills_total",
              "headshot_kills_total",
              "kill_reward_total",
              "alive_time_total",
              "objective_total",
              "mvps",
              "utility_damage_total",
              "3k_rounds_total",
              "4k_rounds_total",
              "ace_rounds_total",
              "enemies_flashed_total",
              "total_rounds_played",
              "team_rounds_total",
              "rounds_played_this_phase",
              "team_name",
              "score",
              "money_saved_total",
              "cash_earned_total",
              "is_connected",
              "ping",
              "game_phase",
              "active_weapon_name",
              "active_weapon_original_owner",
              "weapon_purchases_this_match",
              "weapon_purchases_this_round",
              "has_helmet",
              "armor_value",
              "has_defuser",
              "prev_owner",
              "active_weapon_ammo",
              "total_ammo_left",
              "item_def_idx",
              "item_id_high",
              "item_id_low",
              "inventory_position",
              "is_silencer_on",
              "is_scoped",
              "shots_fired",
              "team_surrendered",
              "is_warmup_period",
              "is_freeze_period",
              "is_terrorist_timeout",
              "is_ct_timeout",
              "round_win_status",
              "round_win_reason",
              "team_score_first_half",
              "team_score_second_half",
              "inventory",
              "inventory_as_ids"]

    fields = [
        "X", "Y", "Z", "health", "score", "mvps", "is_alive", "balance", "inventory",
        "inventory_as_ids", "life_state", "pitch", "yaw", "is_auto_muted", "crosshair_code",
        "pending_team_num", "player_color", "ever_played_on_team", "is_coach_team", "rank",
        "rank_if_win", "rank_if_loss", "rank_if_tie", "comp_wins", "comp_rank_type",
        "is_controlling_bot", "has_controlled_bot_this_round", "can_control_bot", "has_defuser",
        "has_helmet", "spawn_time", "death_time", "game_time", "is_connected", "player_name",
        "player_steamid", "fov", "start_balance", "total_cash_spent", "cash_spent_this_round",
        "music_kit_id", "leader_honors", "teacher_honors", "friendly_honors", "ping",
        "move_collide", "move_type", "team_num", "active_weapon", "looking_at_weapon",
        "holding_look_at_weapon", "next_attack_time", "duck_time_ms", "max_speed",
        "max_fall_velo", "duck_amount", "duck_speed", "duck_overrdie", "old_jump_pressed",
        "jump_until", "jump_velo", "fall_velo", "in_crouch", "crouch_state", "ducked",
        "ducking", "in_duck_jump", "allow_auto_movement", "jump_time_ms", "last_duck_time",
        "is_rescuing", "weapon_purchases_this_match", "weapon_purchases_this_round", "spotted",
        "approximate_spotted_by", "time_last_injury", "direction_last_injury", "player_state",
        "passive_items", "is_scoped", "is_walking", "resume_zoom", "is_defusing",
        "is_grabbing_hostage", "blocking_use_in_progess", "molotov_damage_time",
        "moved_since_spawn", "in_bomb_zone", "in_buy_zone", "in_no_defuse_area",
        "killed_by_taser", "move_state", "which_bomb_zone", "in_hostage_rescue_zone", "stamina",
        "direction", "shots_fired", "armor_value", "velo_modifier",
        "ground_accel_linear_frac_last_time", "flash_duration", "flash_max_alpha",
        "wait_for_no_attack", "last_place_name", "is_strafing", "round_start_equip_value",
        "current_equip_value", "velocity", "velocity_X", "velocity_Y", "velocity_Z",
        "agent_skin", "user_id", "entity_id", "is_airborne", "aim_punch_angle",
        "aim_punch_angle_vel",

        # Button states
        "FORWARD", "LEFT", "RIGHT", "BACK", "FIRE", "RIGHTCLICK", "RELOAD", "INSPECT",
        "USE", "ZOOM", "SCOREBOARD", "WALK", "buttons",

        # Game state
        "team_rounds_total", "team_surrendered", "team_name", "team_score_overtime",
        "team_match_stat", "team_num_map_victories", "team_score_first_half",
        "team_score_second_half", "team_clan_name", "is_freeze_period", "is_warmup_period",
        "warmup_period_end", "warmup_period_start", "is_terrorist_timeout", "is_ct_timeout",
        "terrorist_timeout_remaining", "ct_timeout_remaining", "num_terrorist_timeouts",
        "num_ct_timeouts", "is_technical_timeout", "is_waiting_for_resume", "match_start_time",
        "round_start_time", "restart_round_time", "game_start_time",
        "time_until_next_phase_start", "game_phase", "total_rounds_played",
        "rounds_played_this_phase", "hostages_remaining", "any_hostages_reached",
        "has_bombites", "has_rescue_zone", "has_buy_zone", "is_matchmaking",
        "match_making_mode", "is_valve_dedicated_server", "gungame_prog_weap_ct",
        "gungame_prog_weap_t", "spectator_slot_count", "is_match_started", "n_best_of_maps",
        "is_bomb_dropped", "is_bomb_planted", "round_win_status", "round_win_reason",
        "terrorist_cant_buy", "ct_cant_buy", "ct_losing_streak", "t_losing_streak",
        "survival_start_time", "round_in_progress",

        # Weapon fields
        "active_weapon_name", "active_weapon_skin", "active_weapon_ammo",
        "active_weapon_original_owner", "total_ammo_left", "item_def_idx", "weapon_quality",
        "entity_lvl", "item_id_high", "item_id_low", "item_account_id", "inventory_position",
        "is_initialized", "econ_item_attribute_def_idx", "initial_value", "refundable_currency",
        "set_bonus", "custom_name", "orig_owner_xuid_low", "orig_owner_xuid_high",
        "fall_back_paint_kit", "fall_back_seed", "fall_back_wear", "fall_back_stat_track",
        "m_iState", "fire_seq_start_time", "fire_seq_start_time_change",
        "is_player_fire_event_primary", "weapon_mode", "accuracy_penalty", "i_recoil_idx",
        "fl_recoil_idx", "is_burst_mode", "post_pone_fire_ready_time", "is_in_reload",
        "reload_visually_complete", "dropped_at_time", "is_hauled_back", "is_silencer_on",
        "time_silencer_switch_complete", "orig_team_number", "prev_owner", "last_shot_time",
        "iron_sight_mode", "num_empty_attacks", "zoom_lvl", "burst_shots_remaining",
        "needs_bolt_action", "next_primary_attack_tick", "next_primary_attack_tick_ratio",
        "next_secondary_attack_tick", "next_secondary_attack_tick_ratio", "weapon_float",
        "weapon_paint_seed", "weapon_stickers",

        # User commands
        "usercmd_viewangle_x", "usercmd_viewangle_y", "usercmd_viewangle_z",
        "usercmd_buttonstate_1", "usercmd_buttonstate_2", "usercmd_buttonstate_3",
        "usercmd_consumed_server_angle_changes", "usercmd_forward_move", "usercmd_left_move",
        "usercmd_impulse", "usercmd_mouse_dx", "usercmd_mouse_dy", "usercmd_left_hand_desired",
        "usercmd_weapon_select", "usercmd_input_history",

        # Aggregate stats
        "kills_total", "deaths_total", "assists_total", "alive_time_total",
        "headshot_kills_total", "ace_rounds_total", "4k_rounds_total", "3k_rounds_total",
        "damage_total", "objective_total", "utility_damage_total", "enemies_flashed_total",
        "equipment_value_total", "money_saved_total", "kill_reward_total", "cash_earned_total"]

    # sets up spreadsheet and puts data in spreadsheet
    #round_mvp_df = get_df("ROUND MVP:", round_mvp_ticks)
    player_death_df = get_df("PLAYER DEATH:", player_death_ticks)
    ##bomb_planted_df = get_df("BOMB PLANTED:", bomb_planted_ticks)
    ##bomb_defused_df = get_df("BOMB DEFUSED:", bomb_defused_ticks)
    #hostage_rescued_df = get_df("HOSTAGE RESCUED:", hostage_rescued_ticks)
    weapon_fire_df = get_df("WEAPON FIRE:", weapon_fire_ticks)
    #flashbang_detonate_df = get_df("FLASHBANG DETONATE:", flashbang_detonate_ticks)
    #hegrenade_detonate_df = get_df("HEGRENADE DETONATE:", hegrenade_detonate_ticks)
    #molotov_detonate_df = get_df("MOLOTOV DETONATE:", molotov_detonate_tick)
    #smokegrenade_detonate_df = get_df("SMOKEGRENADE DETONATE:", smokegrenade_detonate_ticks)
    player_hurt_df = get_df("PLAYER HURT:", player_hurt_ticks)
    ##player_blind_df = get_df("PLAYER BLIND:", player_blind_ticks)
    player_connect_df = get_df("PLAYER CONNECT:", player_connect_ticks)
    #player_activate_df = get_df("PLAYER ACTIVATE:", player_activate_ticks)
    begin_new_match_df = get_df("BEGIN NEW:", begin_new_match_ticks)

    round_start_df = get_df("ROUND START:", round_start_ticks)
    round_freeze_end_df = get_df("ROUND FREEZE:", round_freeze_end_ticks)
    round_end_df = get_df("ROUND END:", round_end_ticks)

    # export to excel
    excel_file = os.path.basename(file) + ".xlsx"
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        #round_mvp_df.to_excel(writer, sheet_name="Round MVP", index=False)
        player_death_df.to_excel(writer, sheet_name="Player Death", index=False)
        ##bomb_planted_df.to_excel(writer, sheet_name="Bomb Planted", index=False)
        ##bomb_defused_df.to_excel(writer, sheet_name="Bomb Defused", index=False)
        #hostage_rescued_df.to_excel(writer, sheet_name="Hostage Rescued", index=False)
        weapon_fire_df.to_excel(writer, sheet_name="Weapon Fire", index=False)
        #flashbang_detonate_df.to_excel(writer, sheet_name="Flashbang Detonate", index=False)
        #hegrenade_detonate_df.to_excel(writer, sheet_name="HE Detonate", index=False)
        #molotov_detonate_df.to_excel(writer, sheet_name="Molotov Detonate", index=False)
        #smokegrenade_detonate_df.to_excel(writer, sheet_name="Smoke Detonate", index=False)
        player_hurt_df.to_excel(writer, sheet_name="Player Hurt", index=False)
        ##player_blind_df.to_excel(writer, sheet_name="Player Blind", index=False)
        #player_activate_df.to_excel(writer, sheet_name="Player Activate", index=False)
        player_connect_df.to_excel(writer, sheet_name="Player Connect", index=False)
        begin_new_match_df.to_excel(writer, sheet_name="Begin New", index=False)

        round_start_df.to_excel(writer, sheet_name="Round Start", index=False)
        round_freeze_end_df.to_excel(writer, sheet_name="Round Freeze", index=False)
        round_end_df.to_excel(writer, sheet_name="Round End", index=False)

    # reloads file
    wb = load_workbook(excel_file)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # colors sections
    highlight_excel_sections(wb)

    # saves file
    wb.save(excel_file)
    wb.close()